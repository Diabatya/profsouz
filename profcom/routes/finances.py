import re
from datetime import date
from decimal import Decimal
from io import BytesIO

from flask import (
    Blueprint,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook

from models import FinanceCommission, FinanceExpense, FinanceMonth, FinanceYear, Protocol, db
from utils import login_required, parse_date, parse_decimal

MONTH_NAMES = [
    "январь",
    "февраль",
    "март",
    "апрель",
    "май",
    "июнь",
    "июль",
    "август",
    "сентябрь",
    "октябрь",
    "ноябрь",
    "декабрь",
]

bp = Blueprint("finances", __name__, url_prefix="/finances")


def _is_admin():
    return "admin_id" in session


def _distribute(year, gross):
    gross = Decimal(gross) if gross is not None else Decimal(0)
    return {
        "mpo": round(gross * year.mpo_percent / Decimal(100), 2),
        "opo": round(gross * year.opo_percent / Decimal(100), 2),
        "ppo": round(gross * year.ppo_percent / Decimal(100), 2),
        "charity": round(gross * year.charity_percent / Decimal(100), 2),
    }


def _year_totals(year):
    months = year.months.order_by(FinanceMonth.month).all()
    total_gross = sum((m.gross_amount for m in months), Decimal(0))
    mpo = sum((m.mpo_amount for m in months), Decimal(0))
    opo = sum((m.opo_amount for m in months), Decimal(0))
    ppo_income = sum((m.ppo_amount for m in months), Decimal(0))
    charity_income = sum((m.charity_amount for m in months), Decimal(0))
    expenses_ppo = sum(
        (e.amount for e in year.expenses.filter_by(fund="ppo").all()),
        Decimal(0),
    )
    expenses_charity = sum(
        (e.amount for e in year.expenses.filter_by(fund="charity").all()),
        Decimal(0),
    )
    commissions = sum(
        (c.amount for c in year.commissions.all()),
        Decimal(0),
    )
    ppo_closing = year.ppo_opening + ppo_income - expenses_ppo - commissions
    charity_closing = year.charity_opening + charity_income - expenses_charity
    return {
        "gross": total_gross,
        "mpo": mpo,
        "opo": opo,
        "ppo_income": ppo_income,
        "charity_income": charity_income,
        "expenses_ppo": expenses_ppo,
        "expenses_charity": expenses_charity,
        "commissions": commissions,
        "ppo_closing": ppo_closing,
        "charity_closing": charity_closing,
        "total_closing": ppo_closing + charity_closing,
    }


def _build_month_rows(year, months, expenses):
    fm_map = {m.month: m for m in months}
    expenses_by_month = {}
    for e in expenses:
        expenses_by_month.setdefault(e.date.month, []).append(e)
    cumulative = Decimal(0)
    rows = []
    for i in range(1, 13):
        fm = fm_map.get(i)
        income = fm.gross_amount if fm else Decimal(0)
        month_expenses = expenses_by_month.get(i, [])
        expense_total = sum((e.amount for e in month_expenses), Decimal(0))
        balance = income - expense_total
        cumulative += balance
        rows.append(
            {
                "month": i,
                "name": MONTH_NAMES[i - 1].title(),
                "fm": fm,
                "income": income,
                "expenses": month_expenses,
                "balance": balance,
                "cumulative": cumulative,
            }
        )
    return rows


def _parse_protocol_info(text):
    if not text:
        return None, None
    match = re.search(r"№\s*(\d+)\s*от\s*(\d{2}\.\d{2}\.\d{4})", str(text))
    if match:
        return match.group(1), parse_date(match.group(2))
    return None, None


def _find_protocol(number, protocol_date):
    q = Protocol.query
    if number:
        q = q.filter(Protocol.number.ilike(str(number)))
    if protocol_date:
        q = q.filter(Protocol.date == protocol_date)
    return q.first()


@bp.route("/protocol-info")
@login_required
def protocol_info():
    number = (request.args.get("number") or "").strip()
    if not number:
        return jsonify({"found": False})
    protocol = _find_protocol(number)
    if not protocol:
        return jsonify({"found": False})
    return jsonify(
        {
            "found": True,
            "id": protocol.id,
            "number": protocol.number,
            "date": protocol.date.strftime("%d.%m.%Y") if protocol.date else "",
            "amount": str(protocol.total_amount) if protocol.total_amount else "0",
        }
    )


@bp.route("/")
@login_required
def index():
    years = FinanceYear.query.order_by(FinanceYear.year.desc()).all()
    year_id = request.args.get("year_id", type=int)
    year = None
    if year_id:
        year = db.session.get(FinanceYear, year_id)
    if not year and years:
        year = years[0]
    months = []
    totals = {}
    expenses = []
    commissions = []
    chart_bars = []
    month_rows = []
    if year:
        months = year.months.order_by(FinanceMonth.month).all()
        expenses = year.expenses.order_by(FinanceExpense.date.desc()).all()
        commissions = year.commissions.order_by(FinanceCommission.date.desc()).all()
        totals = _year_totals(year)
        expenses_by_month = {}
        for e in expenses:
            expenses_by_month[e.date.month] = (
                expenses_by_month.get(e.date.month, Decimal(0)) + e.amount
            )
        chart_raw = [
            (
                MONTH_NAMES[m.month - 1].title(),
                float(m.gross_amount),
                float(expenses_by_month.get(m.month, Decimal(0))),
            )
            for m in months
        ]
        max_val = (
            max([i for _, i, _ in chart_raw] + [e for _, _, e in chart_raw]) if chart_raw else 1
        )
        for name, inc, exp in chart_raw:
            chart_bars.append(
                {
                    "name": name,
                    "income_h": int(inc / max_val * 100) if max_val else 0,
                    "expense_h": int(exp / max_val * 100) if max_val else 0,
                }
            )
        month_rows = _build_month_rows(year, months, expenses)
    year_summaries = [(y, _year_totals(y)) for y in years]
    return render_template(
        "finances/index.html",
        years=years,
        year=year,
        year_summaries=year_summaries,
        months=months,
        expenses=expenses,
        commissions=commissions,
        totals=totals,
        chart_bars=chart_bars,
        month_rows=month_rows,
        month_names=MONTH_NAMES,
        is_admin=_is_admin(),
    )


@bp.route("/year/add", methods=["POST"])
@login_required
def year_add():
    if not _is_admin():
        flash("Недостаточно прав", "danger")
        return redirect(url_for("finances.index"))

    year = request.form.get("year", type=int)
    prev_year = FinanceYear.query.filter_by(year=year - 1).first()
    if prev_year:
        prev_totals = _year_totals(prev_year)
        ppo_opening = prev_totals["ppo_closing"]
        charity_opening = prev_totals["charity_closing"]
    else:
        ppo_opening = parse_decimal(request.form.get("ppo_opening", "0"))
        charity_opening = parse_decimal(request.form.get("charity_opening", "0"))
    mpo_percent = parse_decimal(request.form.get("mpo_percent", "15"))
    opo_percent = parse_decimal(request.form.get("opo_percent", "10"))
    ppo_percent = parse_decimal(request.form.get("ppo_percent", "70"))
    charity_percent = parse_decimal(request.form.get("charity_percent", "5"))

    if not year:
        flash("Укажите год", "danger")
        return redirect(url_for("finances.index"))

    if mpo_percent + opo_percent + ppo_percent + charity_percent != Decimal(100):
        flash("Сумма процентов должна быть равна 100", "danger")
        return redirect(url_for("finances.index"))

    if FinanceYear.query.filter_by(year=year).first():
        flash("Такой год уже существует", "danger")
        return redirect(url_for("finances.index"))

    fy = FinanceYear(
        year=year,
        ppo_opening=ppo_opening,
        charity_opening=charity_opening,
        mpo_percent=mpo_percent,
        opo_percent=opo_percent,
        ppo_percent=ppo_percent,
        charity_percent=charity_percent,
    )
    db.session.add(fy)
    db.session.commit()
    flash("Год добавлен", "success")
    return redirect(url_for("finances.index", year_id=fy.id))


@bp.route("/year/<int:id>/edit", methods=["POST"])
@login_required
def year_edit(id):
    if not _is_admin():
        flash("Недостаточно прав", "danger")
        return redirect(url_for("finances.index"))

    fy = db.session.get(FinanceYear, id) or abort(404)
    fy.ppo_opening = parse_decimal(request.form.get("ppo_opening", "0"))
    fy.charity_opening = parse_decimal(request.form.get("charity_opening", "0"))
    fy.mpo_percent = parse_decimal(request.form.get("mpo_percent", "15"))
    fy.opo_percent = parse_decimal(request.form.get("opo_percent", "10"))
    fy.ppo_percent = parse_decimal(request.form.get("ppo_percent", "70"))
    fy.charity_percent = parse_decimal(request.form.get("charity_percent", "5"))
    fy.note = (request.form.get("note") or "").strip() or None

    if fy.mpo_percent + fy.opo_percent + fy.ppo_percent + fy.charity_percent != Decimal(100):
        flash("Сумма процентов должна быть равна 100", "danger")
        return redirect(url_for("finances.index", year_id=fy.id))

    db.session.commit()
    flash("Год обновлен", "success")
    return redirect(url_for("finances.index", year_id=fy.id))


@bp.route("/year/<int:id>/recalc", methods=["POST"])
@login_required
def year_recalc(id):
    if not _is_admin():
        flash("Недостаточно прав", "danger")
        return redirect(url_for("finances.index"))
    fy = db.session.get(FinanceYear, id) or abort(404)
    prev_year = FinanceYear.query.filter_by(year=fy.year - 1).first()
    if not prev_year:
        flash("Предыдущий год не найден", "danger")
        return redirect(url_for("finances.index", year_id=fy.id))
    prev_totals = _year_totals(prev_year)
    fy.ppo_opening = prev_totals["ppo_closing"]
    fy.charity_opening = prev_totals["charity_closing"]
    db.session.commit()
    flash("Остатки пересчитаны с учётом комиссий", "success")
    return redirect(url_for("finances.index", year_id=fy.id))


@bp.route("/month/save", methods=["POST"])
@login_required
def month_save():
    if not _is_admin():
        flash("Недостаточно прав", "danger")
        return redirect(url_for("finances.index"))

    year_id = request.form.get("year_id", type=int)
    month = request.form.get("month", type=int)
    gross = parse_decimal(request.form.get("gross_amount", "0"))
    received = parse_date(request.form.get("date_received", ""))

    year = db.session.get(FinanceYear, year_id) or abort(404)
    if not month or not (1 <= month <= 12):
        flash("Укажите месяц", "danger")
        return redirect(url_for("finances.index", year_id=year.id))

    dist = _distribute(year, gross)
    fm = FinanceMonth.query.filter_by(year_id=year.id, month=month).first()
    if not fm:
        fm = FinanceMonth(year_id=year.id, month=month)
        db.session.add(fm)

    fm.gross_amount = gross
    fm.mpo_amount = dist["mpo"]
    fm.opo_amount = dist["opo"]
    fm.ppo_amount = dist["ppo"]
    fm.charity_amount = dist["charity"]
    fm.date_received = received or date(year.year, month, 1)
    db.session.commit()
    flash("Поступление сохранено", "success")
    return redirect(url_for("finances.index", year_id=year.id))


@bp.route("/expense/add", methods=["POST"])
@login_required
def expense_add():
    if not _is_admin():
        flash("Недостаточно прав", "danger")
        return redirect(url_for("finances.index"))

    year_id = request.form.get("year_id", type=int)
    year = db.session.get(FinanceYear, year_id) or abort(404)
    expense_date = parse_date(request.form.get("date", ""))
    fund = (request.form.get("fund") or "ppo").strip()
    protocol_number = (request.form.get("protocol_number") or "").strip() or None
    protocol_date = parse_date(request.form.get("protocol_date", ""))
    description = (request.form.get("description") or "").strip()
    amount = parse_decimal(request.form.get("amount", "0"))

    if not expense_date or not amount or fund not in ("ppo", "charity"):
        flash("Заполните дату, сумму и фонд", "danger")
        return redirect(url_for("finances.index", year_id=year.id))

    protocol = (
        _find_protocol(protocol_number, protocol_date) if protocol_number or protocol_date else None
    )
    if protocol and not protocol_date:
        protocol_date = protocol.date
    if not description:
        if protocol and protocol.number:
            description = f"Расход по протоколу №{protocol.number}"
            if protocol_date:
                description += f" от {protocol_date.strftime('%d.%m.%Y')}"
        elif protocol_number:
            description = f"Расход по протоколу №{protocol_number}"
        else:
            description = "Расход"

    fe = FinanceExpense(
        year_id=year.id,
        date=expense_date,
        fund=fund,
        protocol_number=protocol_number,
        protocol_date=protocol_date,
        protocol_id=protocol.id if protocol else None,
        description=description,
        amount=amount,
    )
    db.session.add(fe)
    db.session.commit()
    flash("Расход добавлен", "success")
    return redirect(url_for("finances.index", year_id=year.id))


@bp.route("/expense/<int:id>/delete", methods=["POST"])
@login_required
def expense_delete(id):
    if not _is_admin():
        flash("Недостаточно прав", "danger")
        return redirect(url_for("finances.index"))

    fe = db.session.get(FinanceExpense, id) or abort(404)
    year_id = fe.year_id
    db.session.delete(fe)
    db.session.commit()
    flash("Расход удален", "success")
    return redirect(url_for("finances.index", year_id=year_id))


@bp.route("/commission/add", methods=["POST"])
@login_required
def commission_add():
    if not _is_admin():
        flash("Недостаточно прав", "danger")
        return redirect(url_for("finances.index"))

    year_id = request.form.get("year_id", type=int)
    year = db.session.get(FinanceYear, year_id) or abort(404)
    commission_date = parse_date(request.form.get("date", ""))
    amount = parse_decimal(request.form.get("amount", "0"))
    description = (request.form.get("description") or "").strip() or "Банковская комиссия"

    if not commission_date or not amount:
        flash("Заполните дату и сумму комиссии", "danger")
        return redirect(url_for("finances.index", year_id=year.id))

    fc = FinanceCommission(
        year_id=year.id,
        date=commission_date,
        amount=amount,
        description=description,
    )
    db.session.add(fc)
    db.session.commit()
    flash("Комиссия добавлена", "success")
    return redirect(url_for("finances.index", year_id=year.id))


@bp.route("/commission/<int:id>/delete", methods=["POST"])
@login_required
def commission_delete(id):
    if not _is_admin():
        flash("Недостаточно прав", "danger")
        return redirect(url_for("finances.index"))

    fc = db.session.get(FinanceCommission, id) or abort(404)
    year_id = fc.year_id
    db.session.delete(fc)
    db.session.commit()
    flash("Комиссия удалена", "success")
    return redirect(url_for("finances.index", year_id=year_id))


@bp.route("/import", methods=["POST"])
@login_required
def import_excel():
    if not _is_admin():
        flash("Недостаточно прав", "danger")
        return redirect(url_for("finances.index"))

    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("Загрузите файл Excel (.xlsx)", "danger")
        return redirect(url_for("finances.index"))

    wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
    ws = None
    for sheet in wb.worksheets:
        if re.fullmatch(r"\d{4}", str(sheet.title).strip()):
            ws = sheet
            break
    if ws is None:
        ws = wb.worksheets[0]
    sheet_name = str(ws.title).strip()
    year_match = re.fullmatch(r"(\d{4})", sheet_name)
    if not year_match:
        flash("Не удалось определить год из названия листа Excel", "danger")
        return redirect(url_for("finances.index"))
    year_value = int(year_match.group(1))

    def _cell_value(r, c):
        v = ws.cell(row=r, column=c).value
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return Decimal(str(v))
        return parse_decimal(str(v))

    def _cell_text(r, c):
        v = ws.cell(row=r, column=c).value
        return str(v).strip() if v is not None else None

    # Остатки на начало
    ppo_opening = _cell_value(19, 5) or Decimal(0)
    charity_opening = _cell_value(26, 5) or Decimal(0)

    year = FinanceYear.query.filter_by(year=year_value).first()
    if year:
        # очистим старые данные перед перезагрузкой
        for m in year.months.all():
            db.session.delete(m)
        for e in year.expenses.all():
            db.session.delete(e)
        for c in year.commissions.all():
            db.session.delete(c)
    else:
        year = FinanceYear(
            year=year_value,
            ppo_opening=ppo_opening,
            charity_opening=charity_opening,
            mpo_percent=15,
            opo_percent=10,
            ppo_percent=70,
            charity_percent=5,
        )
        db.session.add(year)
    db.session.commit()
    db.session.refresh(year)

    # Поступления по месяцам
    for row_idx, month_name in enumerate(MONTH_NAMES, start=7):
        amount = _cell_value(row_idx, 5)
        if amount and amount > 0:
            month_num = MONTH_NAMES.index(month_name) + 1
            dist = _distribute(year, amount)
            fm = FinanceMonth.query.filter_by(year_id=year.id, month=month_num).first()
            if not fm:
                fm = FinanceMonth(year_id=year.id, month=month_num)
                db.session.add(fm)
            fm.gross_amount = amount
            fm.mpo_amount = dist["mpo"]
            fm.opo_amount = dist["opo"]
            fm.ppo_amount = dist["ppo"]
            fm.charity_amount = dist["charity"]
            fm.date_received = date(year_value, month_num, 1)

    # Банковская комиссия
    commission_total = _cell_value(22, 5)
    if commission_total and commission_total > 0:
        fc = FinanceCommission(
            year_id=year.id,
            date=date(year_value, 12, 31),
            amount=commission_total,
            description=f"Банковская комиссия за {year_value} год (импорт)",
        )
        db.session.add(fc)

    # Расходы ППО (строки 33-50) и благотворительного фонда (54-67)
    def _import_expenses(start_row, end_row, fund):
        for r in range(start_row, end_row + 1):
            text = _cell_text(r, 4)
            amount = _cell_value(r, 5)
            if not text or not amount or amount <= 0:
                continue
            number, pdate = _parse_protocol_info(text)
            protocol = _find_protocol(number, pdate) if number or pdate else None
            fe = FinanceExpense(
                year_id=year.id,
                date=pdate or date(year_value, 1, 1),
                fund=fund,
                protocol_number=number,
                protocol_date=pdate,
                protocol_id=protocol.id if protocol else None,
                description=text,
                amount=amount,
            )
            db.session.add(fe)

    _import_expenses(33, 50, "ppo")
    _import_expenses(54, 67, "charity")

    db.session.commit()
    flash(f"Бюджет {year_value} импортирован", "success")
    return redirect(url_for("finances.index", year_id=year.id))


@bp.route("/export/<int:year_id>")
@login_required
def export(year_id):
    year = db.session.get(FinanceYear, year_id) or abort(404)
    totals = _year_totals(year)
    wb = Workbook()

    ws_income = wb.active
    ws_income.title = "Поступления"
    ws_income.append(["Месяц", "Валовый сбор", "МПО", "ОПО", "ППО", "Благотворительный фонд"])
    for m in year.months.order_by(FinanceMonth.month).all():
        ws_income.append(
            [
                MONTH_NAMES[m.month - 1].title(),
                float(m.gross_amount),
                float(m.mpo_amount),
                float(m.opo_amount),
                float(m.ppo_amount),
                float(m.charity_amount),
            ]
        )
    ws_income.append(
        [
            "Итого",
            float(totals["gross"]),
            float(totals["mpo"]),
            float(totals["opo"]),
            float(totals["ppo_income"]),
            float(totals["charity_income"]),
        ]
    )

    ws_expenses = wb.create_sheet("Расходы")
    ws_expenses.append(["Дата", "Фонд", "Протокол", "Дата протокола", "Описание", "Сумма"])
    for e in year.expenses.order_by(FinanceExpense.date).all():
        ws_expenses.append(
            [
                e.date.isoformat(),
                "ППО" if e.fund == "ppo" else "Благотворительный фонд",
                e.protocol_number or "",
                e.protocol_date.isoformat() if e.protocol_date else "",
                e.description,
                float(e.amount),
            ]
        )

    ws_commissions = wb.create_sheet("Комиссии")
    ws_commissions.append(["Дата", "Сумма", "Описание"])
    for c in year.commissions.order_by(FinanceCommission.date).all():
        ws_commissions.append([c.date.isoformat(), float(c.amount), c.description])

    ws_summary = wb.create_sheet("Сводка")
    ws_summary.append(["Показатель", "Сумма"])
    ws_summary.append(["Остаток ППО на начало года", float(year.ppo_opening)])
    ws_summary.append(["Остаток фонда на начало года", float(year.charity_opening)])
    ws_summary.append(["Всего поступило", float(totals["gross"])])
    ws_summary.append(["Из них МПО", float(totals["mpo"])])
    ws_summary.append(["Из них ОПО", float(totals["opo"])])
    ws_summary.append(["Из них ППО", float(totals["ppo_income"])])
    ws_summary.append(["Из них фонд", float(totals["charity_income"])])
    ws_summary.append(["Расходы ППО", float(totals["expenses_ppo"])])
    ws_summary.append(["Расходы фонда", float(totals["expenses_charity"])])
    ws_summary.append(["Банковская комиссия", float(totals["commissions"])])
    ws_summary.append(["Остаток ППО на конец года", float(totals["ppo_closing"])])
    ws_summary.append(["Остаток фонда на конец года", float(totals["charity_closing"])])
    ws_summary.append(["Остаток всего", float(totals["total_closing"])])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"finances_{year.year}.xlsx",
    )
