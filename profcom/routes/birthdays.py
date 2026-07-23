from datetime import date
from io import BytesIO

from flask import (
    Blueprint,
    abort,
    render_template,
    request,
    send_file,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy import extract

from models import AnniversarySetting, Member, db
from utils import login_required

bp = Blueprint("birthdays", __name__, url_prefix="/birthdays")


def _age_and_amount(member, today):
    if not member.birth_date:
        return None, 0, False, 0
    age = today.year - member.birth_date.year
    if (today.month, today.day) < (member.birth_date.month, member.birth_date.day):
        age -= 1
    setting = db.session.get(AnniversarySetting, age)
    amount = setting.amount if setting else 0
    is_anniversary = setting is not None and amount > 0
    return member.birth_date, age, is_anniversary, amount


@bp.route("/")
@login_required
def index():
    today = date.today()
    month = request.args.get("month", today.month, type=int)
    if not 1 <= month <= 12:
        month = today.month
    members = (
        Member.query.filter(
            Member.status != "excluded",
            Member.birth_date != None,
            extract("month", Member.birth_date) == month,
        )
        .order_by(extract("day", Member.birth_date))
        .all()
    )
    rows = []
    for m in members:
        bdate, age, is_anniversary, amount = _age_and_amount(m, today)
        rows.append({"member": m, "age": age, "is_anniversary": is_anniversary, "amount": amount})
    return render_template(
        "birthdays.html",
        rows=rows,
        month=month,
        months=list(range(1, 13)),
    )


@bp.route("/export")
@login_required
def export():
    today = date.today()
    month = request.args.get("month", today.month, type=int)
    if not 1 <= month <= 12:
        month = today.month
    members = (
        Member.query.filter(
            Member.status != "excluded",
            Member.birth_date != None,
            extract("month", Member.birth_date) == month,
        )
        .order_by(extract("day", Member.birth_date))
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Именинники"
    ws.append(["ФИО", "Отдел", "Дата рождения", "Возраст", "Юбилей", "Сумма"])
    fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    for m in members:
        _, age, is_anniversary, amount = _age_and_amount(m, today)
        cells = [
            m.full_name,
            m.department or "",
            m.birth_date.strftime("%d.%m.%Y") if m.birth_date else "",
            age,
            "Да" if is_anniversary else "Нет",
            float(amount) if is_anniversary else "",
        ]
        ws.append(cells)
        if is_anniversary:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"imeninniki_{month:02d}_{today.year}.xlsx"
    return send_file(
        stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
