import os
import re
import uuid
from datetime import date, datetime
from io import BytesIO

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_from_directory,
    url_for,
)
from openpyxl import load_workbook

from models import Group, Member, MemberChild, MemberComment, MemberStatusHistory, Position, db
from utils import (
    apply_sort,
    dictionary_values,
    login_required,
    parse_date,
    save_dictionary_value,
    title_name,
)

bp = Blueprint("members", __name__, url_prefix="/members")


ALLOWED_PHOTO = {".jpg", ".jpeg", ".png", ".webp"}


def _save_photo(file):
    if not file or not file.filename:
        return None
    ext = os.path.splitext(file.filename.lower())[1]
    if ext not in ALLOWED_PHOTO:
        return None
    photo_dir = os.path.join(current_app.config["UPLOAD_FOLDER"], "photos")
    os.makedirs(photo_dir, exist_ok=True)
    try:
        from PIL import Image

        img = Image.open(file.stream)
        img = img.convert("RGB")
        width, height = img.size
        size = min(width, height)
        left = (width - size) // 2
        top = (height - size) // 2
        img = img.crop((left, top, left + size, top + size))
        img = img.resize((400, 400), getattr(Image, "Resampling", Image).LANCZOS)
        filename = f"{uuid.uuid4().hex}.jpg"
        path = os.path.join(photo_dir, filename)
        img.save(path, "JPEG", quality=85, optimize=True)
        return f"photos/{filename}"
    except Exception:
        filename = f"{uuid.uuid4().hex}{ext}"
        path = os.path.join(photo_dir, filename)
        file.stream.seek(0)
        file.save(path)
        return f"photos/{filename}"


def _delete_photo(member):
    if member.photo_path:
        full = os.path.join(current_app.config["UPLOAD_FOLDER"], member.photo_path)
        try:
            if os.path.exists(full):
                os.remove(full)
        except OSError:
            pass
        member.photo_path = None


@bp.route("/")
@login_required
def index():
    q = Member.query
    dept = request.args.get("dept", "")
    status = request.args.get("status", "")
    gender = request.args.get("gender", "")
    position = request.args.get("position", "")
    search = request.args.get("search", "")
    sort = request.args.get("sort", "full_name")
    order = request.args.get("order", "asc")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    view = request.args.get("view", "table")

    if status:
        q = q.filter(Member.status == status)
    else:
        q = q.filter(Member.status == "active")

    if dept:
        q = q.filter(Member.department == dept)
    if gender:
        q = q.filter(Member.gender == gender)
    if position:
        q = q.filter(Member.position == position)
    if search:
        q = q.filter(Member.full_name.ilike(f"%{search}%"))

    q = apply_sort(
        q, sort, order, Member, ["full_name", "department", "position", "birth_date", "entry_date"]
    )
    pagination = q.paginate(page=page, per_page=max(per_page, 5), error_out=False)
    departments = [
        r[0]
        for r in db.session.query(Member.department).distinct().order_by(Member.department).all()
    ]
    positions = [
        r[0]
        for r in db.session.query(Member.position).distinct().order_by(Member.position).all()
        if r[0]
    ]
    return render_template(
        "members/list.html",
        members=pagination.items,
        pagination=pagination,
        departments=departments,
        positions=positions,
        department_choices=dictionary_values("department"),
        position_choices=dictionary_values("position"),
        selected_dept=dept,
        selected_status=status,
        selected_position=position,
        selected_gender=gender,
        search=search,
        sort=sort,
        order=order,
        per_page=per_page,
        view=view,
    )


@bp.route("/add", methods=["GET", "POST"])
@login_required
def add():
    from forms import MemberForm

    form = MemberForm()
    form.organization_position_id.choices = _position_choices()
    if form.validate_on_submit():
        full_name = title_name(form.full_name.data)
        department = (
            save_dictionary_value("department", form.department.data.strip())
            or form.department.data.strip()
        )
        position = (form.position.data or "").strip() or None
        position = save_dictionary_value("position", position) or position
        organization_position_id = form.organization_position_id.data or None
        phone = (form.phone.data or "").strip() or None
        gender = form.gender.data
        birth_date = parse_date(form.birth_date.data)
        entry_date = parse_date(form.entry_date.data)

        if not birth_date:
            flash("Укажите дату рождения", "danger")
            return render_template(
                "members/add.html",
                form=form,
                department_choices=dictionary_values("department"),
                position_choices=dictionary_values("position"),
            )

        member = Member(
            full_name=full_name,
            department=department,
            position=position,
            organization_position_id=organization_position_id,
            phone=phone,
            birth_date=birth_date,
            entry_date=entry_date,
            status="active",
        )
        if gender in ("male", "female"):
            member.gender = gender
        else:
            member.gender = Member.detect_gender(full_name)
        db.session.add(member)
        db.session.flush()
        photo_path = _save_photo(request.files.get("photo"))
        if photo_path:
            member.photo_path = photo_path
        db.session.add(
            MemberStatusHistory(
                member_id=member.id, old_status=None, new_status="active", note="Создание"
            )
        )
        db.session.commit()
        flash("Член профсоюза добавлен", "success")
        return redirect(url_for("members.index"))
    return render_template(
        "members/add.html",
        form=form,
        department_choices=dictionary_values("department"),
        position_choices=dictionary_values("position"),
    )


def _position_choices():
    positions = (
        Position.query.filter_by(active=True, scope="organization")
        .order_by(Position.level, Position.name)
        .all()
    )
    return [(0, "-")] + [(p.id, f"{p.name} (ур. {p.level})") for p in positions]


def _normalize_full_name(name):
    name = str(name or "").strip()
    name = re.sub(r"\s+", " ", name)
    return name.lower().replace("ё", "е")


def _find_existing_member(full_name, existing_map):
    norm = _normalize_full_name(full_name)
    return existing_map.get(norm) if norm else None


def _parse_date_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    if not s:
        return None
    return parse_date(s)


def _is_truthy_cell(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    return str(value).strip().lower() in {"1", "да", "yes", "+", "true"}


def _get_or_create_group(name, group_type="category"):
    group = Group.query.filter_by(name=name, type=group_type).first()
    if not group:
        group = Group(name=name, type=group_type)
        db.session.add(group)
        db.session.flush()
    return group


def _parse_import_preview(file_storage):
    wb = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
    ws = wb.active
    first_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip() if h else "" for h in first_row]

    header_aliases = {
        "ФИО": "full_name",
        "Отдел": "department",
        "Подразделение": "department",
        "Должность": "position",
        "Должность в профсоюзе": "union_position",
        "Пол": "gender",
        "Дата рождения": "birth_date",
        "Дата вступления": "entry_date",
    }
    special_cols = {
        "М": "male",
        "Ж": "female",
        "Декрет": "maternity",
        "МОП": "mop",
    }
    col_map = {}
    special_map = {}
    for i, h in enumerate(headers):
        h = str(h).strip()
        if h in header_aliases:
            col_map[header_aliases[h]] = i
        if h in special_cols:
            special_map[special_cols[h]] = i

    required = ["full_name", "department", "birth_date"]
    if not all(k in col_map for k in required):
        return None, "В файле должны быть колонки: ФИО, Подразделение/Отдел, Дата рождения"

    preview = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        full_name_raw = row[col_map["full_name"]] if "full_name" in col_map else None
        full_name = title_name(str(full_name_raw).strip()) if full_name_raw else ""
        department_raw = row[col_map["department"]] if "department" in col_map else None
        department = str(department_raw).strip() if department_raw else ""
        birth_date = (
            _parse_date_value(row[col_map["birth_date"]]) if "birth_date" in col_map else None
        )
        entry_date = (
            _parse_date_value(row[col_map["entry_date"]]) if "entry_date" in col_map else None
        )
        position = (
            str(row[col_map["position"]]).strip()
            if "position" in col_map and row[col_map["position"]]
            else ""
        )
        union_position = (
            str(row[col_map["union_position"]]).strip()
            if "union_position" in col_map and row[col_map["union_position"]]
            else ""
        )
        gender = ""
        if "gender" in col_map and row[col_map["gender"]]:
            g = str(row[col_map["gender"]]).strip().lower()
            if g in ("м", "муж", "мужской", "male"):
                gender = "male"
            elif g in ("ж", "жен", "женский", "female"):
                gender = "female"
        else:
            if "male" in special_map and _is_truthy_cell(row[special_map["male"]]):
                gender = "male"
            elif "female" in special_map and _is_truthy_cell(row[special_map["female"]]):
                gender = "female"

        maternity = "maternity" in special_map and _is_truthy_cell(row[special_map["maternity"]])
        mop = "mop" in special_map and _is_truthy_cell(row[special_map["mop"]])

        errors = []
        if not full_name:
            errors.append("ФИО")
        if not department:
            errors.append("отдел")
        if not birth_date:
            errors.append("дата рождения")

        preview.append(
            {
                "row_idx": row_idx,
                "full_name": full_name,
                "department": department,
                "position": position,
                "union_position": union_position,
                "gender": gender,
                "birth_date": birth_date.strftime("%Y-%m-%d") if birth_date else "",
                "entry_date": entry_date.strftime("%Y-%m-%d") if entry_date else "",
                "maternity": maternity,
                "mop": mop,
                "errors": errors,
                "valid": not errors,
            }
        )
    return preview, None


@bp.route("/import", methods=["GET", "POST"])
@login_required
def import_members():
    if request.method == "POST":
        file = request.files.get("file")
        preview = None
        error = None
        if file:
            if not file.filename.lower().endswith(".xlsx"):
                flash("Загрузите файл Excel (.xlsx)", "danger")
                return redirect(url_for("members.import_members"))
            preview, error = _parse_import_preview(file)
            if error:
                flash(error, "danger")
                return redirect(url_for("members.import_members"))

        action = request.form.get("action") or "save"
        if action == "save":
            if preview is not None:
                rows = preview
            else:
                row_count = request.form.get("row_count", type=int) or 0
                rows = []
                for i in range(row_count):
                    rows.append(
                        {
                            "full_name": request.form.get(f"full_name_{i}", ""),
                            "department": request.form.get(f"department_{i}", ""),
                            "position": request.form.get(f"position_{i}", ""),
                            "union_position": request.form.get(f"union_position_{i}", ""),
                            "gender": request.form.get(f"gender_{i}", ""),
                            "birth_date": request.form.get(f"birth_date_{i}", ""),
                            "entry_date": request.form.get(f"entry_date_{i}", ""),
                            "maternity": request.form.get(f"maternity_{i}") == "1",
                            "mop": request.form.get(f"mop_{i}") == "1",
                        }
                    )

            created = 0
            updated = 0
            errors = []
            existing_map = {_normalize_full_name(m.full_name): m for m in Member.query.all()}
            seen_departments = set()
            seen_positions = set()
            category_groups = {
                "maternity": _get_or_create_group("Декрет"),
                "mop": _get_or_create_group("МОП"),
            }
            seen_names = set()
            try:
                for idx, row in enumerate(rows, start=1):
                    full_name = row["full_name"].strip()
                    norm = _normalize_full_name(full_name)
                    if norm in seen_names:
                        errors.append(f"Строка {idx}: дубликат ФИО")
                        continue
                    seen_names.add(norm)
                    department = row["department"].strip()
                    position = row["position"].strip() or None
                    union_position_name = row["union_position"].strip()
                    gender = row["gender"]
                    birth_date = _parse_date_value(row["birth_date"])
                    entry_date = _parse_date_value(row["entry_date"])
                    if not full_name or not department or not birth_date:
                        errors.append(f"Строка {idx}: не заполнены обязательные поля")
                        continue

                    member = _find_existing_member(full_name, existing_map)
                    organization_position_id = None
                    if union_position_name:
                        pos = Position.query.filter_by(
                            name=union_position_name, scope="organization"
                        ).first()
                        if not pos:
                            pos = Position(name=union_position_name, scope="organization")
                            db.session.add(pos)
                            db.session.flush()
                        organization_position_id = pos.id

                    if member:
                        member.department = department
                        member.position = position
                        member.organization_position_id = organization_position_id
                        member.birth_date = birth_date
                        member.entry_date = entry_date
                        member.gender = gender
                        updated += 1
                    else:
                        member = Member(
                            full_name=full_name,
                            department=department,
                            position=position,
                            organization_position_id=organization_position_id,
                            birth_date=birth_date,
                            entry_date=entry_date,
                            status="active",
                            gender=gender,
                        )
                        db.session.add(member)
                        db.session.flush()
                        existing_map[_normalize_full_name(full_name)] = member
                        created += 1

                    for key, group in category_groups.items():
                        is_set = bool(row[key])
                        if is_set and group not in member.groups:
                            member.groups.append(group)
                        elif not is_set and group in member.groups:
                            member.groups.remove(group)

                    seen_departments.add(department)
                    if position:
                        seen_positions.add(position)
                    db.session.add(
                        MemberStatusHistory(
                            member_id=member.id,
                            old_status=None,
                            new_status="active",
                            note="Импорт из Excel",
                        )
                    )
                db.session.commit()
                for value in seen_departments:
                    save_dictionary_value("department", value)
                for value in seen_positions:
                    save_dictionary_value("position", value)
            except Exception as e:
                db.session.rollback()
                flash(f"Ошибка сохранения: {e}", "danger")
                return redirect(url_for("members.import_members"))

            msg = f"Создано {created}, обновлено {updated} членов"
            if errors:
                msg += f", ошибок: {len(errors)}"
                for e in errors[:10]:
                    flash(e, "danger")
                if len(errors) > 10:
                    flash(f"... и ещё {len(errors) - 10} ошибок", "danger")
            flash(msg, "success")
            return redirect(url_for("members.index"))

        return render_template("members/import.html", preview=preview)
    return render_template("members/import.html")


@bp.route("/import/template")
@login_required
def import_template():
    from utils import excel_response

    headers = [
        "№",
        "Кол-во",
        "Код",
        "Подразделение",
        "ФИО",
        "М",
        "Ж",
        "Декрет",
        "МОП",
        "Дата рождения",
    ]
    example = [1, 1, "00", "Руководство", "Иванов Иван Иванович", 1, "", "", "", date(1985, 3, 15)]
    return excel_response(headers, [example], "shablon_importa.xlsx")


@bp.route("/import/children", methods=["GET", "POST"])
@login_required
def import_children():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.lower().endswith(".xlsx"):
            flash("Загрузите файл Excel (.xlsx)", "danger")
            return redirect(url_for("members.import_children"))

        wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
        ws = wb.active
        first_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        headers = [str(h).strip() if h else "" for h in first_row]

        header_aliases = {
            "ФИО": "child_name",
            "Ребенок": "child_name",
            "ФИО ребенка": "child_name",
            "Дата рождения": "birth_date",
            "День рождения": "birth_date",
            "Родитель": "parent_name",
            "ФИО родителя": "parent_name",
            "ФИО члена": "parent_name",
            "Член профсоюза": "parent_name",
        }
        col_map = {}
        for i, h in enumerate(headers):
            h = str(h).strip()
            if h in header_aliases:
                col_map[header_aliases[h]] = i

        if "child_name" not in col_map or "parent_name" not in col_map:
            flash(
                "В файле должны быть колонки: ФИО ребенка и ФИО родителя (члена профсоюза)",
                "danger",
            )
            return redirect(url_for("members.import_children"))

        parents_map = {_normalize_full_name(m.full_name): m for m in Member.query.all()}
        created = 0
        updated = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            child_raw = row[col_map["child_name"]] if col_map["child_name"] < len(row) else None
            child_name = title_name(str(child_raw).strip()) if child_raw else ""
            parent_raw = row[col_map["parent_name"]] if col_map["parent_name"] < len(row) else None
            parent_name = title_name(str(parent_raw).strip()) if parent_raw else ""
            birth_date = (
                _parse_date_value(row[col_map["birth_date"]])
                if "birth_date" in col_map and col_map["birth_date"] < len(row)
                else None
            )

            if not child_name or not parent_name:
                errors.append(f"Строка {row_idx}: не указано ФИО ребенка или родителя")
                continue

            parent = parents_map.get(_normalize_full_name(parent_name))
            if not parent:
                errors.append(
                    f"Строка {row_idx}: ребёнок '{child_name}' — "
                    f"родитель '{parent_name}' не найден; исправьте вручную"
                )
                continue

            child = MemberChild.query.filter_by(member_id=parent.id, full_name=child_name).first()
            if child:
                child.birth_date = birth_date
                updated += 1
            else:
                db.session.add(
                    MemberChild(member_id=parent.id, full_name=child_name, birth_date=birth_date)
                )
                created += 1

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Ошибка импорта: {e}", "danger")
            return redirect(url_for("members.import_children"))

        msg = f"Создано {created}, обновлено {updated} детей"
        if errors:
            msg += f", ошибок: {len(errors)}"
            for e in errors[:10]:
                flash(e, "danger")
            if len(errors) > 10:
                flash(f"... и ещё {len(errors) - 10} ошибок", "danger")
        flash(msg, "success")
        return redirect(url_for("members.index"))

    return render_template("members/import_children.html")


@bp.route("/import/children/template")
@login_required
def import_children_template():
    from utils import excel_response

    headers = ["ФИО ребенка", "Дата рождения", "ФИО родителя"]
    example = ["Иванов Петр Иванович", "2015-09-10", "Иванов Иван Иванович"]
    return excel_response(headers, [example], "shablon_deti.xlsx")


@bp.route("/<int:id>")
@login_required
def detail(id):
    member = db.session.get(Member, id) or abort(404)
    return render_template("members/detail.html", member=member)


@bp.route("/<int:id>/photo")
def photo(id):
    member = db.session.get(Member, id) or abort(404)
    if member.photo_path:
        return send_from_directory(current_app.config["UPLOAD_FOLDER"], member.photo_path)
    abort(404)


@bp.route("/<int:id>/edit", methods=["GET", "POST"])
@login_required
def edit(id):
    from forms import MemberForm

    member = db.session.get(Member, id) or abort(404)
    form = MemberForm(obj=member)
    form.organization_position_id.choices = _position_choices()
    if member.organization_position_id:
        form.organization_position_id.data = member.organization_position_id
    if form.validate_on_submit():
        full_name = title_name(form.full_name.data)
        department = (
            save_dictionary_value("department", form.department.data.strip())
            or form.department.data.strip()
        )
        position = (form.position.data or "").strip() or None
        position = save_dictionary_value("position", position) or position
        organization_position_id = form.organization_position_id.data or None
        phone = (form.phone.data or "").strip() or None
        gender = form.gender.data
        birth_date = parse_date(form.birth_date.data)
        entry_date = parse_date(form.entry_date.data)

        if not birth_date:
            flash("Укажите дату рождения", "danger")
            return render_template(
                "members/edit.html",
                member=member,
                form=form,
                department_choices=dictionary_values("department"),
                position_choices=dictionary_values("position"),
            )

        member.full_name = full_name
        member.department = department
        member.position = position
        member.organization_position_id = organization_position_id
        member.phone = phone
        member.birth_date = birth_date
        member.entry_date = entry_date
        if gender in ("male", "female"):
            member.gender = gender
        else:
            member.gender = Member.detect_gender(full_name)

        if request.form.get("delete_photo"):
            _delete_photo(member)
        else:
            photo_path = _save_photo(request.files.get("photo"))
            if photo_path:
                _delete_photo(member)
                member.photo_path = photo_path

        db.session.commit()
        flash("Данные обновлены", "success")
        return redirect(url_for("members.detail", id=member.id))
    return render_template(
        "members/edit.html",
        member=member,
        form=form,
        department_choices=dictionary_values("department"),
        position_choices=dictionary_values("position"),
    )


@bp.route("/<int:id>/delete", methods=["POST"])
@login_required
def delete(id):
    member = db.session.get(Member, id) or abort(404)
    if member.status == "not_member":
        flash("Член уже исключён", "warning")
        return redirect(url_for("members.index"))

    old_status = member.status
    member.status = "not_member"
    db.session.add(
        MemberStatusHistory(
            member_id=member.id,
            old_status=old_status,
            new_status="not_member",
            exit_date=date.today(),
            note="Исключение",
        )
    )
    db.session.commit()
    flash("Член исключён", "success")
    return redirect(url_for("members.index"))


@bp.route("/<int:id>/restore", methods=["POST"])
@login_required
def restore(id):
    member = db.session.get(Member, id) or abort(404)
    if member.status == "active":
        flash("Член уже активен", "warning")
        return redirect(url_for("members.index"))

    old_status = member.status
    member.status = "active"
    db.session.add(
        MemberStatusHistory(
            member_id=member.id, old_status=old_status, new_status="active", note="Восстановление"
        )
    )
    db.session.commit()
    flash("Член восстановлен", "success")
    return redirect(url_for("members.index"))


@bp.route("/bulk", methods=["POST"])
@login_required
def bulk():
    ids_str = request.form.get("member_ids", "")
    action = request.form.get("action", "")
    value = request.form.get("value", "").strip()
    ids = [int(i) for i in ids_str.split(",") if i.isdigit()]
    members = Member.query.filter(Member.id.in_(ids)).all()
    count = 0
    for member in members:
        if action == "change_department" and value:
            member.department = value
            count += 1
        elif action == "change_position":
            member.position = value or None
            count += 1
        elif action == "change_gender" and value in ("male", "female"):
            member.gender = value
            count += 1
        elif action == "exclude" and member.status == "active":
            old_status = member.status
            member.status = "not_member"
            db.session.add(
                MemberStatusHistory(
                    member_id=member.id,
                    old_status=old_status,
                    new_status="not_member",
                    exit_date=date.today(),
                    note="Массовое исключение",
                )
            )
            count += 1
    db.session.commit()
    flash(f"Обработано {count} членов", "success")
    return redirect(url_for("members.index"))


@bp.route("/profkom")
@login_required
def profkom():
    gender = request.args.get("gender", "")
    groups = Group.query.filter_by(type="profkom").order_by(Group.name).all()
    rank = {"Председатель": 0, "Заместитель председателя": 1, "Секретарь": 2, "Член профкома": 3}
    members_by_group = {}
    for g in groups:
        members = g.members
        if gender in ("male", "female"):
            members = [m for m in members if m.gender_or_detect == gender]
        members_by_group[g.id] = sorted(
            members, key=lambda m: rank.get(m.position or "Член профкома", 99)
        )
    return render_template(
        "members/profkom.html",
        groups=groups,
        members_by_group=members_by_group,
        selected_gender=gender,
    )


@bp.route("/<int:member_id>/children/add", methods=["POST"])
@login_required
def add_child(member_id):
    member = db.session.get(Member, member_id) or abort(404)
    full_name = title_name((request.form.get("full_name") or "").strip())
    birth_date = parse_date(request.form.get("birth_date"))
    gender = request.form.get("gender", "").strip()
    if not full_name:
        flash("Укажите ФИО ребенка", "danger")
        return redirect(url_for("members.detail", id=member.id))
    child = MemberChild(
        member_id=member.id,
        full_name=full_name,
        birth_date=birth_date,
        gender=gender if gender in ("male", "female") else None,
    )
    db.session.add(child)
    db.session.commit()
    flash("Ребенок добавлен", "success")
    return redirect(url_for("members.detail", id=member.id))


@bp.route("/children/<int:child_id>/delete", methods=["POST"])
@login_required
def delete_child(child_id):
    child = db.session.get(MemberChild, child_id) or abort(404)
    member_id = child.member_id
    db.session.delete(child)
    db.session.commit()
    flash("Запись удалена", "success")
    return redirect(url_for("members.detail", id=member_id))


@bp.route("/<int:member_id>/comment", methods=["POST"])
@login_required
def add_comment(member_id):
    member = db.session.get(Member, member_id) or abort(404)
    text = (request.form.get("text") or "").strip()
    if not text:
        flash("Введите текст комментария", "danger")
        return redirect(url_for("members.detail", id=member.id))
    db.session.add(MemberComment(member_id=member.id, text=text))
    db.session.commit()
    flash("Комментарий добавлен", "success")
    return redirect(url_for("members.detail", id=member.id))


@bp.route("/comments/<int:comment_id>/delete", methods=["POST"])
@login_required
def delete_comment(comment_id):
    comment = db.session.get(MemberComment, comment_id) or abort(404)
    member_id = comment.member_id
    db.session.delete(comment)
    db.session.commit()
    flash("Комментарий удалён", "success")
    return redirect(url_for("members.detail", id=member_id))


@bp.route("/gifts")
@login_required
def gifts():
    department = request.args.get("department", "").strip()
    position = request.args.get("position", "").strip()
    min_age = request.args.get("min_age", type=int)
    max_age = request.args.get("max_age", type=int)
    child_gender = request.args.get("child_gender", "").strip().lower()
    status = request.args.get("status", "active").strip().lower()
    export = request.args.get("export", "")

    q = Member.query
    if status:
        q = q.filter(Member.status == status)
    if department:
        q = q.filter(Member.department == department)
    if position:
        q = q.filter(Member.position == position)
    members = q.order_by(Member.full_name).all()

    today = date.today()
    result = []
    total_adults = 0
    total_children = 0
    total_gifts = 0
    for m in members:
        children = []
        for child in m.children:
            age = None
            if child.birth_date:
                age = (
                    today.year
                    - child.birth_date.year
                    - ((today.month, today.day) < (child.birth_date.month, child.birth_date.day))
                )
            if min_age is not None and (age is None or age < min_age):
                continue
            if max_age is not None and (age is None or age > max_age):
                continue
            if child_gender and child.gender != child_gender:
                continue
            children.append({"child": child, "age": age})
        total_adults += 1
        total_children += len(children)
        member_total = 1 + len(children)
        total_gifts += member_total
        result.append({"member": m, "children": children, "total": member_total})

    if export:
        from utils import excel_response

        rows = []
        for r in result:
            child_info = ", ".join(
                f"{c['child'].full_name} ({c['age']})"
                for c in r["children"]
                if c["age"] is not None
            )
            rows.append(
                [
                    r["member"].full_name,
                    r["member"].department,
                    r["member"].position or "-",
                    len(r["children"]),
                    r["total"],
                    child_info,
                ]
            )
        headers = ["ФИО", "Отдел", "Должность", "Детей", "Всего подарков", "Дети (возраст)"]
        return excel_response(headers, rows, "presents.xlsx")

    return render_template(
        "members/gifts.html",
        rows=result,
        department=department,
        position=position,
        min_age=min_age,
        max_age=max_age,
        child_gender=child_gender,
        status=status,
        total_adults=total_adults,
        total_children=total_children,
        total_gifts=total_gifts,
        departments=dictionary_values("department"),
        positions=dictionary_values("position"),
    )


@bp.route("/committees")
@login_required
def committees():
    groups = Group.query.filter_by(type="committee").order_by(Group.name).all()
    return render_template("members/committees.html", groups=groups)


@bp.route("/groups", methods=["GET", "POST"])
@login_required
def groups():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        gtype = request.form.get("type", "other")
        if name:
            db.session.add(Group(name=name, type=gtype))
            db.session.commit()
            flash("Группа создана", "success")
        else:
            flash("Укажите название группы", "danger")
        return redirect(url_for("members.groups"))
    q = Group.query
    sort = request.args.get("sort", "name")
    order = request.args.get("order", "asc")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    q = apply_sort(q, sort, order, Group, ["name", "type"])
    pagination = q.paginate(page=page, per_page=max(per_page, 5), error_out=False)
    return render_template(
        "members/groups.html",
        groups=pagination.items,
        pagination=pagination,
        sort=sort,
        order=order,
        per_page=per_page,
    )


@bp.route("/groups/<int:id>/delete", methods=["POST"])
@login_required
def delete_group(id):
    group = db.session.get(Group, id) or abort(404)
    group.members = []
    db.session.delete(group)
    db.session.commit()
    flash("Группа удалена", "success")
    return redirect(url_for("members.groups"))


@bp.route("/groups/<int:id>", methods=["GET", "POST"])
@login_required
def group_detail(id):
    group = db.session.get(Group, id) or abort(404)
    if request.method == "POST":
        member_id = request.form.get("member_id", type=int)
        if member_id:
            member = db.session.get(Member, member_id)
            if member and member not in group.members:
                group.members.append(member)
                db.session.commit()
                flash("Член добавлен в группу", "success")
    available = Member.query.filter_by(status="active").order_by(Member.full_name).all()
    return render_template("members/group_detail.html", group=group, available=available)


@bp.route("/groups/<int:group_id>/remove/<int:member_id>", methods=["POST"])
@login_required
def remove_from_group(group_id, member_id):
    group = db.session.get(Group, group_id) or abort(404)
    member = db.session.get(Member, member_id) or abort(404)
    if member in group.members:
        group.members.remove(member)
        db.session.commit()
        flash("Член исключён из группы", "success")
    return redirect(url_for("members.group_detail", id=group.id))
